#coding:utf-8
from numpy import *
import numpy as np
import logging  
import time  
import os  
import jieba  
import glob  
import random  
import copy  
import chardet  
import gensim  
import jieba.analyse 
import jieba.posseg as pseg
import codecs
import json
import re
from gensim import corpora,similarities, models  
from pprint import pprint  
from sklearn import feature_extraction  
from sklearn.feature_extraction.text import TfidfTransformer  
from sklearn.feature_extraction.text import CountVectorizer  
from sklearn.decomposition import PCA  
from sklearn.cluster import SpectralClustering
from openpyxl import load_workbook
from gensim import corpora, models, similarities


#--------------------------------------------------------------------

stop_words = 'stop_words_eng.txt'
stopwords = codecs.open(stop_words,'r').readlines()
stopwords = [ w.strip() for w in stopwords ]
#print(stopwords);
stop_flag = ['x', 'c', 'u','d', 'p', 't', 'uj', 'm', 'f', 'r']
corpus = []
res = {}
res['links'] = []
res['nodes'] = []

#--------------------------------------------------------------------

def tokenization(text):
    result = []
    words = pseg.cut(text)
    for word, flag in words:
        if flag not in stop_flag and word not in stopwords:
            result.append(word)
    return result

#--------------------------------------------------------------------

def deleteSymbol(text):
	punc = '[,.!\']'
	newText = re.sub(punc,'',text);
	# print(newText);
	return newText;
	
#--------------------------------------------------------------------
#读取xlsx文件
def read_from_xlsx(filename):
    workbook = load_workbook(filename);
    sheets = workbook.sheetnames;
    booksheet = workbook[sheets[0]];
    rows = booksheet.rows;
    line = [];
    for row in rows:
        line.append([col.value for col in row]);
    data = [];
    f = open("abstract.txt","w");
    for i in range(len(line)):
        if i == 0:
            continue;
        curLine = [];
        for j in range(len(line[i])):
            if j == 11:
            	curLine = tokenization(str(line[i][j]));
            	for k in range(len(curLine)):
            		f.write(str(curLine[k]) + " ");
            	#print(curLine);
        f.write("\n");
        data.append(curLine);
    #print(data);
    f.close();
    return data;

#--------------------------------------------------------------------

datas = read_from_xlsx("case1_data.xlsx");