#coding:utf-8
from numpy import *
import numpy as np
import logging  
import time  
import os  
import jieba  
import glob  
import random  
import copy  
import chardet  
import gensim  
import jieba.analyse 
import jieba.posseg as pseg
import codecs
import json

from os.path import join as pjoin
from gensim import corpora,similarities, models  
from pprint import pprint  
from sklearn import feature_extraction  
from sklearn.feature_extraction.text import TfidfTransformer  
from sklearn.feature_extraction.text import CountVectorizer  
from sklearn.decomposition import PCA  
from sklearn.cluster import SpectralClustering
from openpyxl import load_workbook
from gensim import corpora, models, similarities

#--------------------------------------------------------------------
#读取xlsx文件
def read_from_xlsx(filename):
    workbook = load_workbook(filename);
    sheets = workbook.sheetnames;
    booksheet = workbook[sheets[0]];
    rows = booksheet.rows;
    line = [];
    for row in rows:
        line.append([col.value for col in row]);
    data = [];
    for i in range(len(line)):
        if i == 0:
            continue;
        curLine = [];
        for j in range(len(line[i])):
            if j == 0 or j == 2 or j == 4 or j == 6 or j == 17 or j == 18:
            	if(j == 17 and line[i][j] == None):
            		continue;
            	curLine.append(line[i][j])
        data.append(curLine);
    #print(data);
    return data;

#--------------------------------------------------------------------
res = {};
res['keywords'] = [];
res['force'] = [];
def concentrate_keyword(data):
#将关键词提炼出来
	filename = "data_test_kw.json";
	fr = open(filename,"a");
	f = open("keywords.txt","w");
	author = [];
	keyword = [];
	authors = [];
	years = [];
	books = [];
	allkeyword = {"a","b"};
	allkeyword.clear();
	cnt = 0;
	for i in range(len(data)):
		curLine = "";
		curLine_author = "";
		curLine_b = "";
		curLine_y = "";
		for j in range(len(data[i])):
			if j == 0:
				author.append(data[i][j]);
			elif j == 4 or j == 5:
				curLine = curLine + str(data[i][j]);
			elif j == 1:
				curLine_author = str(data[i][j]);
			elif j == 2:
				curLine_y = str(data[i][j]);
			elif j == 3:
				curLine_b = str(data[i][j]);
		# res['keywords'].append({
		# 	"id": str(author[cnt]),
		# 	"keyword": str(curLine),
		# });
		curLine = str(curLine).split(";");
		curLine_author = str(curLine_author).split(";");
		authors.append(curLine_author);
		years.append(curLine_y);
		books.append(curLine_b);
		keyword.append(curLine);
		allkeyword.update(curLine);
		cnt = cnt + 1;
		#print(author[i]);
		#print(keyword[i]);
	#print(allkeyword);
	for i in range(len(keyword)):
		keyword_ab = {"a","b"};
		keyword_ab.clear();
		for j in range(len(keyword)):
			# if i == j:
			# 	continue;
			keyword_ab.update(keyword[i]);
			keyword_ab.update(keyword[j]);
			cnt = 0;
			for a in range(len(keyword[i])):
				for b in range(len(keyword[j])):
					if keyword[i][a] == keyword[j][b]:
						cnt = cnt + 1;
			print(str(i) +" " + str(j) + " " + str(len(keyword_ab)));
			print(cnt/len(keyword_ab));
			
			v3 = 0;
			if years[i] == years[j]:
				v3 = v3 + 0.5;
			if books[i] == books[j]:
				v3 = v3 + 0.5;
			res['force'].append({
				"source": str(i),
				"target": str(j),
				"value_kw": cnt/len(keyword_ab),
				"value_at": v3
			});
			keyword_ab.clear();
			print(i*42 + j);

	# for word in allkeyword:
	# 	f.write(str(word));
	# 	f.write(";");
	#print(res);
	f.close();
	#fr.write(json.dumps(res,indent=2))  
	fr.close();
	return;
#--------------------------------------------------------------------


keyword_author = read_from_xlsx("case1_data.xlsx");
concentrate_keyword(keyword_author);
#add_force(keyword_author);